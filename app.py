from flask import Flask, request, send_from_directory, redirect, make_response
from openpyxl import Workbook, load_workbook
import os
import datetime
from zipfile import ZipFile

app = Flask(__name__)
ADMIN_PASSWORD = "password123"

# Create storage folders
if not os.path.exists("voice_reports"):
    os.makedirs("voice_reports")

# Create Excel file if missing
if not os.path.exists("reports.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Engineer", "Date", "Audio File"])
    wb.save("reports.xlsx")

@app.route("/")
def home():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Engineer Voice Report</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            body { 
                font-family: Arial, sans-serif; 
                margin: 20px; 
                max-width: 600px;
                margin-left: auto;
                margin-right: auto;
            }
            button { 
                padding: 10px 20px; 
                margin: 5px; 
                font-size: 16px; 
                cursor: pointer;
                border: none;
                border-radius: 5px;
            }
            button:disabled { opacity: 0.5; cursor: not-allowed; }
            input[type="text"] { 
                padding: 10px; 
                width: 100%; 
                max-width: 400px;
                font-size: 16px;
                border: 2px solid #ddd;
                border-radius: 5px;
                box-sizing: border-box;
            }
            #status { color: #007bff; font-weight: bold; margin-top: 10px; }
            .recording { color: #dc3545; }
            #timer { font-size: 20px; color: #dc3545; margin: 10px 0; }
            #startBtn { background: #28a745; color: white; }
            #stopBtn { background: #dc3545; color: white; }
            @media (max-width: 600px) {
                body { margin: 15px; }
                button { width: 100%; margin: 5px 0; }
                input[type="text"] { width: 100%; }
            }
        </style>
    </head>
    <body>
        <h2>Engineer Voice Report</h2>
        <label>Engineer Name:</label><br>
        <input type="text" id="name" required><br><br>
        
        <label>Record Voice:</label><br>
        <button onclick="startRecording()" id="startBtn">🎙 Start Recording</button>
        <button onclick="stopRecording()" disabled id="stopBtn">⏹ Stop Recording</button>
        <div id="timer"></div>
        <p id="status"></p>
        
        <form id="reportForm" method="POST" action="/submit" enctype="multipart/form-data" style="display:none;">
            <input type="hidden" name="name" id="hiddenName">
            <input type="file" name="audio" id="audioFile">
        </form>
        
        <script>
        let mediaRecorder;
        let audioChunks = [];
        let timerInterval;
        let seconds = 0;
        
        function updateTimer() {
            seconds++;
            const mins = Math.floor(seconds / 60);
            const secs = seconds % 60;
            document.getElementById("timer").innerHTML = 
                `Recording: ${mins}:${secs.toString().padStart(2, '0')}`;
        }
        
        async function startRecording() {
            const nameInput = document.getElementById("name");
            if (!nameInput.value.trim()) {
                alert("Please enter your name first!");
                return;
            }
            
            try {
                document.getElementById("startBtn").disabled = true;
                document.getElementById("stopBtn").disabled = false;
                document.getElementById("status").innerHTML = "🔴 Recording started...";
                document.getElementById("status").className = "recording";
                
                seconds = 0;
                timerInterval = setInterval(updateTimer, 1000);
                
                // Request audio with specific constraints for better quality
                let stream = await navigator.mediaDevices.getUserMedia({ 
                    audio: {
                        echoCancellation: true,
                        noiseSuppression: true,
                        autoGainControl: true,
                        sampleRate: 48000
                    } 
                });
                
                // Try different MIME types for better compatibility
                let options = { audioBitsPerSecond: 128000 };
                let mimeType = '';
                
                // Try MP4 first (best compatibility)
                if (MediaRecorder.isTypeSupported('audio/mp4')) {
                    mimeType = 'audio/mp4';
                    options.mimeType = mimeType;
                }
                // Try WebM with Opus codec
                else if (MediaRecrecorder.isTypeSupported('audio/webm;codecs=opus')) {
                    mimeType = 'audio/webm;codecs=opus';
                    options.mimeType = mimeType;
                }
                // Fallback to default WebM
                else {
                    mimeType = 'audio/webm';
                    options.mimeType = mimeType;
                }
                
                console.log('Using MIME type:', mimeType);
                
                mediaRecorder = new MediaRecorder(stream, options);
                audioChunks = [];
                
                mediaRecorder.ondataavailable = e => {
                    if (e.data.size > 0) {
                        audioChunks.push(e.data);
                        console.log('Audio chunk received:', e.data.size, 'bytes');
                    }
                };
                
                mediaRecorder.onerror = e => {
                    console.error('MediaRecorder error:', e);
                    document.getElementById("status").innerHTML = "❌ Recording error: " + e.error;
                };
                
                // Start recording with timeslice to get data more frequently
                mediaRecorder.start(1000);
                
            } catch (err) {
                document.getElementById("status").innerHTML = "❌ Error: " + err.message;
                document.getElementById("startBtn").disabled = false;
                document.getElementById("stopBtn").disabled = true;
                clearInterval(timerInterval);
                console.error('Error starting recording:', err);
            }
        }
        
        function stopRecording() {
            if (!mediaRecorder || mediaRecorder.state === 'inactive') {
                console.log('MediaRecorder not active');
                return;
            }
            
            clearInterval(timerInterval);
            document.getElementById("timer").innerHTML = "";
            
            mediaRecorder.stop();
            document.getElementById("status").innerHTML = "⏳ Processing audio...";
            document.getElementById("stopBtn").disabled = true;
            
            mediaRecorder.onstop = async () => {
                console.log('Recording stopped, processing', audioChunks.length, 'chunks');
                
                // Stop all audio tracks
                mediaRecorder.stream.getTracks().forEach(track => {
                    track.stop();
                    console.log('Track stopped:', track.kind);
                });
                
                // Create blob from recorded chunks
                const mimeType = mediaRecorder.mimeType;
                const blob = new Blob(audioChunks, { type: mimeType });
                
                console.log("Final blob size:", blob.size, "bytes");
                console.log("MIME type:", mimeType);
                
                if (blob.size === 0) {
                    document.getElementById("status").innerHTML = "❌ Error: Recording is empty. Please try again.";
                    document.getElementById("startBtn").disabled = false;
                    return;
                }
                
                // Determine file extension based on MIME type
                let extension = 'webm';
                if (mimeType.includes('mp4')) {
                    extension = 'mp4';
                } else if (mimeType.includes('ogg')) {
                    extension = 'ogg';
                }
                
                const file = new File([blob], `recording.${extension}`, { type: mimeType });
                
                const dataTransfer = new DataTransfer();
                dataTransfer.items.add(file);
                document.getElementById("audioFile").files = dataTransfer.files;
                
                // Set name
                document.getElementById("hiddenName").value = document.getElementById("name").value;
                
                // Submit form automatically
                document.getElementById("status").innerHTML = "⏳ Uploading...";
                document.getElementById("reportForm").submit();
            };
        }
        </script>
    </body>
    </html>
    """

@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name")
    audio = request.files.get("audio")
    
    if audio and name:
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Get the file extension from uploaded file
            original_filename = audio.filename
            extension = original_filename.split('.')[-1] if '.' in original_filename else 'webm'
            
            # Save the audio file directly
            filename = f"{name}_{timestamp}.{extension}"
            filepath = os.path.join("voice_reports", filename)
            audio.save(filepath)
            
            file_size = os.path.getsize(filepath)
            print(f"Saved audio file: {filename}, size: {file_size} bytes")
            
            # Save to Excel
            wb = load_workbook("reports.xlsx")
            ws = wb.active
            date_formatted = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([name, date_formatted, filename])
            wb.save("reports.xlsx")
            
            return """
            <html>
            <head>
                <meta http-equiv="refresh" content="2;url=/" />
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; text-align: center; }
                </style>
            </head>
            <body>
                <h3>✅ Report submitted successfully!</h3>
                <p>Your voice recording has been saved.</p>
                <p>Redirecting back to home page...</p>
                <a href="/">Submit another report</a>
            </body>
            </html>
            """
        except Exception as e:
            print(f"Error processing audio: {str(e)}")
            import traceback
            traceback.print_exc()
            return f"<h3>❌ Error processing audio: {str(e)}</h3><a href='/'>Go back</a>"
    else:
        return "<h3>❌ Error: Missing name or audio file</h3><a href='/'>Go back</a>"

# -------- ADMIN PROTECTION --------
def admin_protected(request):
    pwd = request.cookies.get("admin_pass")
    return pwd == ADMIN_PASSWORD

@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "GET":
        return """
        <html>
        <head>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; }
                input { padding: 10px; font-size: 14px; }
                button { padding: 10px 20px; font-size: 14px; cursor: pointer; }
            </style>
        </head>
        <body>
            <h2>Admin Login</h2>
            <form method="POST">
                <input name="password" type="password" placeholder="Enter Password" required>
                <button type="submit">Login</button>
            </form>
        </body>
        </html>
        """
    
    password = request.form.get("password")
    if password == ADMIN_PASSWORD:
        resp = make_response(redirect("/admin"))
        resp.set_cookie("admin_pass", ADMIN_PASSWORD, max_age=86400)  # 24 hours
        return resp
    else:
        return "<h3>❌ Wrong Password</h3><a href='/admin_login'>Try again</a>"

@app.route("/admin")
def admin():
    if not admin_protected(request):
        return redirect("/admin_login")
    
    files = sorted(os.listdir("voice_reports"), reverse=True)
    
    page = """
    <html>
    <head>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            a { text-decoration: none; color: #007bff; }
            a:hover { text-decoration: underline; }
            .file-list { margin-top: 20px; }
            .file-item { 
                padding: 12px; 
                border-bottom: 1px solid #eee; 
                display: flex;
                justify-content: space-between;
                align-items: center;
            }
            .logout { float: right; color: #dc3545; }
            .play-btn { 
                background: #28a745; 
                color: white; 
                padding: 5px 10px; 
                border: none; 
                border-radius: 4px;
                cursor: pointer;
                margin-left: 10px;
            }
        </style>
    </head>
    <body>
        <h2>Admin Panel <a href='/admin_logout' class='logout'>Logout</a></h2>
        <p>
            <a href='/download_excel'>📄 Download Excel Report</a><br><br>
            <a href='/download_all_audio'>🎧 Download All Audio (ZIP)</a><br><br>
        </p>
        
        <h3>Individual Audio Files (""" + str(len(files)) + """ files):</h3>
        <div class="file-list">
    """
    
    for f in files:
        page += f"""
        <div class='file-item'>
            <span>🎵 {f}</span>
            <div>
                <button class='play-btn' onclick="playAudio('/audio/{f}')">▶ Play</button>
                <a href='/audio/{f}' download style='margin-left:10px'>⬇ Download</a>
            </div>
        </div>
        """
    
    page += """
        </div>
        
        <audio id="audioPlayer" controls style="display:none; margin-top:20px; width:100%;"></audio>
        
        <script>
        function playAudio(url) {
            const player = document.getElementById('audioPlayer');
            player.src = url;
            player.style.display = 'block';
            player.play();
        }
        </script>
    </body>
    </html>
    """
    
    return page

@app.route("/admin_logout")
def admin_logout():
    resp = make_response(redirect("/admin_login"))
    resp.set_cookie("admin_pass", "", expires=0)
    return resp

@app.route("/download_excel")
def download_excel():
    if not admin_protected(request):
        return redirect("/admin_login")
    return send_from_directory(".", "reports.xlsx", as_attachment=True)

@app.route("/audio/<filename>")
def download_audio(filename):
    if not admin_protected(request):
        return redirect("/admin_login")
    return send_from_directory("voice_reports", filename)

@app.route("/download_all_audio")
def download_all_audio():
    if not admin_protected(request):
        return redirect("/admin_login")
    
    zip_path = "all_audio.zip"
    with ZipFile(zip_path, "w") as zipf:
        for file in os.listdir("voice_reports"):
            zipf.write(os.path.join("voice_reports", file), file)
    
    return send_from_directory(".", zip_path, as_attachment=True)

# -------- PORT FOR RENDER --------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=False)
